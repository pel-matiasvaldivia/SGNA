# -*- coding: utf-8 -*-
"""Genera:
 - n8n-workflow-auditorias-en-linea.json  (workflow importable en n8n)
 - calendario-n8n.csv                      (fuente para Google Sheets)
 - README-n8n.md                           (instrucciones de instalación)"""
import os, csv, json, uuid
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "plan-marketing-auditorias-en-linea.xlsx")
CSV  = os.path.join(HERE, "calendario-n8n.csv")
WF   = os.path.join(HERE, "n8n-workflow-auditorias-en-linea.json")
README = os.path.join(HERE, "README-n8n.md")

# ------------------------------------------------ 1) CSV para Google Sheets
wb = load_workbook(XLSX)
cal = wb["Calendario 30 días"]
headers = ["Dia","Fecha","DiaSemana","Redes","Pilar","Formato","Titular",
           "Copy","Hashtags","CTA","IdeaVisual","Estado","FechaPublicacion"]
with open(CSV, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(headers)
    for r in range(4, 34):
        row = [cal.cell(r, cc).value for cc in range(1, 12)]
        if row[0] is None:
            continue
        w.writerow(row + ["", ""])   # Estado, FechaPublicacion vacíos
print("OK CSV ->", CSV)

# ------------------------------------------------ 2) Workflow n8n
def nid(): return uuid.uuid4().hex[:16]

n_trigger = nid(); n_read = nid(); n_code = nid(); n_if = nid()
n_li = nid(); n_fb = nid(); n_update = nid(); n_mail = nid()
n_note = nid(); n_note2 = nid()

CODE = r"""// Selecciona el post de HOY y arma el mensaje final.
const rows = $input.all();
const t = new Date();
const dd = String(t.getDate()).padStart(2, '0');
const mm = String(t.getMonth() + 1).padStart(2, '0');
const yyyy = t.getFullYear();
const hoy = `${dd}/${mm}/${yyyy}`;

const match = rows.find(r =>
  String(r.json.Fecha).trim() === hoy &&
  String(r.json.Estado || '').toLowerCase() !== 'publicado'
);

if (!match) { return []; }  // no hay nada para publicar hoy

const j = match.json;
const mensaje = `${j.Copy}\n\n${j.Hashtags}\n\n${j.CTA}`;

return [{
  json: {
    ...j,
    mensaje,
    hoy,
    rowNumber: match.json.row_number,
  }
}];
"""

def node(id_, name, typ, ver, pos, params, extra=None):
    d = {"parameters": params, "id": id_, "name": name, "type": typ,
         "typeVersion": ver, "position": pos}
    if extra: d.update(extra)
    return d

nodes = [
    node(n_trigger, "Cada día 09:00", "n8n-nodes-base.scheduleTrigger", 1.2,
         [-40, 300],
         {"rule": {"interval": [{"field": "cronExpression", "expression": "0 9 * * *"}]}}),

    node(n_read, "Leer calendario (Sheets)", "n8n-nodes-base.googleSheets", 4.5,
         [200, 300],
         {"operation": "read", "documentId": {"__rl": True, "value": "TU_GOOGLE_SHEET_ID", "mode": "id"},
          "sheetName": {"__rl": True, "value": "Calendario", "mode": "name"},
          "options": {}}),

    node(n_code, "Elegir post de hoy", "n8n-nodes-base.code", 2,
         [440, 300],
         {"jsCode": CODE}),

    node(n_if, "¿Hay post para hoy?", "n8n-nodes-base.if", 2.2,
         [680, 300],
         {"conditions": {"options": {"caseSensitive": True, "typeValidation": "loose"},
           "combinator": "and",
           "conditions": [{"id": nid(),
             "leftValue": "={{ $json.mensaje }}", "rightValue": "",
             "operator": {"type": "string", "operation": "exists", "singleValue": True}}]}}),

    node(n_li, "Publicar en LinkedIn", "n8n-nodes-base.linkedIn", 1,
         [940, 180],
         {"postAs": "organization",
          "organization": "={{ $vars.LINKEDIN_ORG_ID }}",
          "text": "={{ $('Elegir post de hoy').item.json.mensaje }}",
          "shareMediaCategory": "NONE", "additionalFields": {}}),

    node(n_fb, "Publicar en Facebook", "n8n-nodes-base.facebookGraphApi", 1,
         [940, 360],
         {"httpRequestMethod": "POST", "graphApiVersion": "v19.0",
          "node": "={{ $vars.FB_PAGE_ID }}", "edge": "feed",
          "options": {"queryParameters": {"parameter": [
              {"name": "message", "value": "={{ $('Elegir post de hoy').item.json.mensaje }}"}]}}}),

    node(n_update, "Marcar como Publicado", "n8n-nodes-base.googleSheets", 4.5,
         [1200, 300],
         {"operation": "update",
          "documentId": {"__rl": True, "value": "TU_GOOGLE_SHEET_ID", "mode": "id"},
          "sheetName": {"__rl": True, "value": "Calendario", "mode": "name"},
          "columns": {"mappingMode": "defineBelow",
            "value": {
              "row_number": "={{ $('Elegir post de hoy').item.json.rowNumber }}",
              "Estado": "Publicado",
              "FechaPublicacion": "={{ $now.format('dd/MM/yyyy HH:mm') }}"},
            "matchingColumns": ["row_number"]},
          "options": {}}),

    node(n_mail, "Avisar al equipo (email)", "n8n-nodes-base.gmail", 2.1,
         [1440, 300],
         {"sendTo": "ventas@auditoriasenlinea.com.ar",
          "subject": "=✅ Publicado hoy: {{ $('Elegir post de hoy').item.json.Titular }}",
          "message": "=Se publicó el post del plan de contenidos.\n\nPilar: {{ $('Elegir post de hoy').item.json.Pilar }}\nRedes: {{ $('Elegir post de hoy').item.json.Redes }}\n\n---\n{{ $('Elegir post de hoy').item.json.mensaje }}\n\nIdea visual: {{ $('Elegir post de hoy').item.json.IdeaVisual }}",
          "options": {}}),

    node(n_note, "", "n8n-nodes-base.stickyNote", 1, [-80, -180],
         {"content": "## Plan de contenidos · Auditorías en Línea\n\n**Qué hace:** cada día a las 09:00 busca el post de hoy en Google Sheets, lo publica en LinkedIn y Facebook, marca la fila como *Publicado* y avisa por email.\n\n**Antes de activar:**\n1. Importá `calendario-n8n.csv` en una Google Sheet (pestaña **Calendario**).\n2. Pegá el ID de la hoja en los 2 nodos de Sheets (`TU_GOOGLE_SHEET_ID`).\n3. Configurá credenciales: Google Sheets, LinkedIn, Facebook y Gmail.\n4. Definí variables (Settings → Variables): `LINKEDIN_ORG_ID`, `FB_PAGE_ID`.", "height": 320, "width": 460}),

    node(n_note2, "", "n8n-nodes-base.stickyNote", 1, [900, -60],
         {"content": "### Aprobación (opcional)\nSi preferís revisar antes de publicar, insertá aquí un nodo de **Telegram/Slack** con botones o un **Wait** de aprobación antes de los nodos de publicación.", "height": 200, "width": 320, "color": 3}),
]

connections = {
    "Cada día 09:00": {"main": [[{"node": "Leer calendario (Sheets)", "type": "main", "index": 0}]]},
    "Leer calendario (Sheets)": {"main": [[{"node": "Elegir post de hoy", "type": "main", "index": 0}]]},
    "Elegir post de hoy": {"main": [[{"node": "¿Hay post para hoy?", "type": "main", "index": 0}]]},
    "¿Hay post para hoy?": {"main": [
        [{"node": "Publicar en LinkedIn", "type": "main", "index": 0},
         {"node": "Publicar en Facebook", "type": "main", "index": 0}],
        []]},
    "Publicar en LinkedIn": {"main": [[{"node": "Marcar como Publicado", "type": "main", "index": 0}]]},
    "Publicar en Facebook": {"main": [[{"node": "Marcar como Publicado", "type": "main", "index": 0}]]},
    "Marcar como Publicado": {"main": [[{"node": "Avisar al equipo (email)", "type": "main", "index": 0}]]},
}

workflow = {
    "name": "Auditorías en Línea · Auto-publicación redes",
    "nodes": nodes,
    "connections": connections,
    "active": False,
    "settings": {"executionOrder": "v1"},
    "pinData": {},
    "meta": {"templateCredsSetupCompleted": False},
}

with open(WF, "w", encoding="utf-8") as f:
    json.dump(workflow, f, ensure_ascii=False, indent=2)
print("OK WF ->", WF)

# ------------------------------------------------ 3) README
readme = """# Auto-gestión del plan de marketing con n8n

Este paquete automatiza la publicación diaria del plan de contenidos de
**Auditorías en Línea** en redes sociales, usando [n8n](https://n8n.io).

## Archivos
- `n8n-workflow-auditorias-en-linea.json` — workflow importable en n8n.
- `calendario-n8n.csv` — el calendario de 30 días, listo para subir a Google Sheets.
- `plan-marketing-auditorias-en-linea.xlsx` — el plan completo (estrategia + calendario).
- `piezas/` — las 30 piezas gráficas (1080×1080) para cada post.

## Cómo funciona
```
Cada día 09:00  →  Leer calendario (Google Sheets)  →  Elegir post de hoy
   →  ¿Hay post?  →  Publicar en LinkedIn + Facebook
   →  Marcar la fila como "Publicado"  →  Avisar al equipo por email
```
El nodo **Elegir post de hoy** busca la fila cuya `Fecha` = hoy (dd/mm/aaaa) y
cuyo `Estado` ≠ *Publicado*. Si no hay, el flujo no publica nada.

## Instalación (paso a paso)
1. **Google Sheets**
   - Creá una hoja nueva e importá `calendario-n8n.csv`.
   - Renombrá la pestaña a **Calendario** (así la busca el workflow).
   - Copiá el **ID de la hoja** (está en la URL, entre `/d/` y `/edit`).

2. **Importar el workflow**
   - En n8n: *Workflows → Import from File* → elegí el `.json`.

3. **Pegar el ID de la hoja**
   - Abrí los nodos **Leer calendario** y **Marcar como Publicado** y reemplazá
     `TU_GOOGLE_SHEET_ID` por el ID real.

4. **Credenciales** (Credentials → New)
   - **Google Sheets** (OAuth2) — para leer/escribir el calendario.
   - **LinkedIn** (OAuth2) — cuenta con permiso de publicar en la página de empresa.
   - **Facebook Graph API** — token de la página con permiso `pages_manage_posts`.
   - **Gmail** (o reemplazá por SMTP) — para el aviso al equipo.

5. **Variables** (Settings → Variables)
   - `LINKEDIN_ORG_ID` → ID numérico de la página de empresa en LinkedIn.
   - `FB_PAGE_ID` → ID de la página de Facebook.

6. **Probar y activar**
   - Ejecutá manualmente (*Execute Workflow*) para ver un post de prueba.
   - Cuando funcione, activá el workflow (toggle **Active**).

## Notas
- **Instagram**: la API exige una imagen alojada públicamente. Para automatizarlo,
  subí las piezas de `piezas/` a un bucket/URL y agregá un nodo *Facebook Graph API*
  con `edge: media` + `media_publish`. (No incluido por defecto.)
- **Aprobación previa**: si querés revisar antes de publicar, insertá un nodo de
  Telegram/Slack o un *Wait* de aprobación antes de los nodos de publicación
  (ver la nota amarilla en el canvas).
- **Horario**: se publica 09:00 (cron `0 9 * * *`). Cambialo en el nodo *Cada día 09:00*.
- **Zona horaria**: configurala en *Settings → Timezone* del workflow (America/Argentina/Mendoza).
- **Datos [dato]**: reemplazá los marcadores `[dato: ...]` del calendario por cifras
  propias antes de automatizar.
"""
with open(README, "w", encoding="utf-8") as f:
    f.write(readme)
print("OK README ->", README)
